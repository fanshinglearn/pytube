# 2025/01/13 pytube 15.0.0 無法使用

import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from pytube import Playlist

import my_function as mf

def confirm(playlist_url, object_list):
    if playlist_url:
        playlist = Playlist(playlist_url)
        videos = mf.PytubeFunction.playlist_to_video_objects(playlist)
        print(f'\n撥放清單名稱： {playlist.title}\n')
        print(f'共 {playlist.length} 部影片，有 {playlist.length - len(playlist.video_urls)} 部無法播放的影片。')
        print(videos[0].title)
        answer = input(f'確定要下載 {len(playlist.video_urls)} 部影片的資訊嗎？ (y/n) ')
    elif object_list:
        videos = object_list
        answer = input(f'確定要下載 {len(object_list)} 部影片的資訊嗎？ (y/n) ')
    else:
        raise ValueError('請提供 Youtube撥放清單網址 或 轉換為Youtube物件的list')
    return answer, videos


def excel_header_format(ws):
    # 左右置中 + 粗體
    for col in range(1, 9):
        char = get_column_letter(col)
        ws[f'{char}1'].alignment = Alignment(horizontal='center')
        ws[f'{char}1'].font = Font(bold=True, size=20)

        # 欄寬
        # 'B
        if col == 2:
            ws.column_dimensions[char].width = 100
        # 'C' 'D' 'E' 'F' 'G' 'H'
        elif col >= 3:
            ws.column_dimensions[char].width = 25

    # 凍結窗格
    ws.freeze_panes = 'B2'


def excel_content_format(ws, row):
    # 上下置中
    for col in range(2, 9):
        char = get_column_letter(col)

        # 'B' 'C'
        if col >= 2 and col <= 3:
            ws[f'{char}{row}'].font = Font(size=16)
            ws[f'{char}{row}'].alignment = Alignment(vertical='center', wrapText=True)
        # 'D' 'E'
        elif col >= 4 and col <= 5:
            ws[f'{char}{row}'].font = Font(size=20)
            ws[f'{char}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        # 'F' 'G' 'H'
        else:
            ws[f'{char}{row}'].font = Font(size=20)
            ws[f'{char}{row}'].alignment = Alignment(vertical='center')


def playlist_to_excel(playlist_url = None, exc_folder_path = 'pytube_excel', object_list = None):

    # 確認
    answer, videos = confirm(playlist_url, object_list)
    
    if answer.lower() == 'y' or answer == '':

        # 創建並切換資料夾
        mf.MyFunction.mkdir_and_chdir_folder(exc_folder_path)

        # 創建 excel
        wb = Workbook()
        ws = wb.active

        # 標題
        ws.append(['縮圖', '標題', '頻道', '觀看次數', '影片長度', '影片網址', '頻道網址', '縮圖網址'])

        # 標題格式
        excel_header_format(ws)

        # 影片資訊
        for row, video in enumerate(videos, start=2):

            # 影片縮圖
            img_url = video.thumbnail_url.split('?')[0]
            img_filename = f'{row}.jpg'
            mf.MyFunction.download_img(img_url, img_filename)
            mf.OpenpyxlFunction.add_thumbnail_to_excel(ws, img_filename, f'A{row}')      

            # 影片資訊
            views = f'{video.views // 10000} 萬' if video.views >= 10000 else video.views
            length = f'{video.length // 60} 分 {video.length % 60} 秒' if video.length >= 60 else f'{video.length} 秒'

            if video.title:
                print(video.title)
            else:
                print('無')
            
            ws.append([None, video.title, video.author, views, length,
                       video.watch_url, video.channel_url, img_url])

            # 超連結
            ws[f'B{row}'].hyperlink = video.watch_url
            ws[f'C{row}'].hyperlink = video.channel_url

            # 內容格式
            excel_content_format(ws, row)

        # 存檔
        exc_path = 'output.xlsx'
        wb.save(exc_path)

        # 開啟 excel
        os.startfile(exc_path)

        # 刪除下載的照片
        mf.MyFunction.delete_images_in_folder()

        print('_' * 30)
        print(f'共 {len(videos)} 部影片')

    else:
        print('取消下載 Q A Q')

if __name__ == '__main__':

    url = 'https://www.youtube.com/playlist?list=PLSCgthA1Anif1w6mKM3O6xlBGGypXtrtN'
    playlist_to_excel(url)


    # 檢查兩個撥放清單是否有相同影片
    # # S 稍後觀看 歌
    # playlist1 = 'https://youtube.com/playlist?list=PLNE0caHGV5eljgnZjHxsBCLCBcTBhB6Qu'

    # # J 日文歌
    # playlist2 = 'https://youtube.com/playlist?list=PLc4D0un5flmPM5Cyd1xmyxW4qbFkf1FHO'

    # # playlist1 = input('輸入第一個播放清單： ')
    # # playlist2 = input('輸入第二個播放清單： ')

    # common_videos = mf.PytubeFunction.common_video_objects(playlist1, playlist2)
    # if common_videos:
    #     print(f'共 {len(common_videos)} 部相同的影片')
    #     playlist_to_excel(object_list=common_videos)
    # else:
    #     print('沒有相同的影片')
